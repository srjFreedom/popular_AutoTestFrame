# -*- coding:utf8 -*-
import os,sys,xlrd
from openpyxl import load_workbook,worksheet

class Transcode():
    def transcode(self, num):
        path = os.path.join(os.path.dirname(sys.argv[0]), 'config', 'error_code.proto')
        with open(path, 'r') as f:
            for each in f.readlines():
                each = each.replace(' ', '')
                p = r'=%s;' % num
                if each.find(p) > -1:
                    return u"结果码意思：" + each.split(r'//')[1].decode('utf-8')
            else:
                return u'没有找到对应的结果码'

class Transexcel():

    def __init__(self):
        pass

    def readexcel(self,xlsName,sheetName):
        xlsdata = []
        filepath = os.path.join(os.path.dirname(sys.argv[0]), 'config', xlsName)
        try:
            wb = load_workbook(filepath,data_only=True)
        except:
            raise BaseException,u"加载配置文件失败"
        try:
            sh = wb[sheetName]
        except:
            raise BaseException,u"加载excel表单失败"
        maxrows = sh.max_row
        maxcols = sh.max_column
        for i in range(1,maxrows+1):
            data_row = []
            for j in range(1,maxcols+1):
                if sh.cell(row=i,column=j).value is None:
                    data_row.append("")
                else:
                    data_row.append(sh.cell(row=i,column=j).value)
            xlsdata.append(data_row)
        return xlsdata

if __name__ == "__main__":
    obj = Transexcel()
    rb = obj.readexcel("Goods.xlsx",u"Sheet1")
    for each in rb:
        print each[1]

